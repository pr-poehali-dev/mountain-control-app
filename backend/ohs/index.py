import json
import os
import base64
from datetime import datetime, date as date_type
import psycopg2

def get_db():
    return psycopg2.connect(os.environ['DATABASE_URL'])

def serialize_default(obj):
    if isinstance(obj, datetime):
        if obj.tzinfo is None:
            return obj.isoformat() + '+00:00'
        return obj.isoformat()
    if isinstance(obj, (date_type,)):
        return obj.isoformat()
    return str(obj)

def json_response(status, body):
    return {
        'statusCode': status,
        'headers': {
            'Content-Type': 'application/json',
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Methods': 'GET, POST, PUT, DELETE, OPTIONS',
            'Access-Control-Allow-Headers': 'Content-Type, Authorization, X-Authorization'
        },
        'body': json.dumps(body, ensure_ascii=False, default=serialize_default)
    }

def col_letter(col_idx):
    result = ''
    while col_idx >= 0:
        result = chr(65 + col_idx % 26) + result
        col_idx = col_idx // 26 - 1
    return result

def parse_excel(file_data):
    from openpyxl import load_workbook
    from io import BytesIO

    raw = base64.b64decode(file_data)
    wb = load_workbook(BytesIO(raw), data_only=False)
    wb_values = load_workbook(BytesIO(raw), data_only=True)

    sheets_result = []

    for idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        ws_vals = wb_values[sheet_name]

        headers = []
        rows_data = []
        formulas = {}
        merged = []

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        if max_row == 0 or max_col == 0:
            sheets_result.append({
                'sheet_name': sheet_name,
                'sheet_index': idx,
                'headers': [],
                'rows_data': [],
                'formulas': {},
                'merged_cells': [],
                'row_count': 0,
                'col_count': 0,
                'column_widths': {}
            })
            continue

        col_widths = {}
        for c in range(1, max_col + 1):
            letter = col_letter(c - 1)
            dim = ws.column_dimensions.get(letter)
            if dim and dim.width:
                col_widths[str(c - 1)] = dim.width

        first_row = None
        for r in range(1, min(max_row + 1, 6)):
            row_vals = []
            for c in range(1, max_col + 1):
                cell = ws_vals.cell(row=r, column=c)
                val = cell.value
                if val is not None:
                    row_vals.append(str(val))
                else:
                    row_vals.append('')
            if any(v.strip() for v in row_vals):
                first_row = r
                headers = row_vals
                break

        if first_row is None:
            first_row = 1
            headers = [col_letter(i) for i in range(max_col)]

        for r in range(first_row + 1, max_row + 1):
            row = []
            for c in range(1, max_col + 1):
                cell_formula = ws.cell(row=r, column=c)
                cell_val = ws_vals.cell(row=r, column=c)

                val = cell_val.value
                if val is None:
                    val = ''
                elif isinstance(val, (datetime, date_type)):
                    val = val.isoformat()
                else:
                    val = str(val) if not isinstance(val, (int, float)) else val

                if isinstance(cell_formula.value, str) and cell_formula.value.startswith('='):
                    cell_ref = '%s%d' % (col_letter(c - 1), r)
                    formulas[cell_ref] = cell_formula.value

                row.append(val)

            if any(str(v).strip() for v in row):
                rows_data.append(row)

        for mg in ws.merged_cells.ranges:
            merged.append(str(mg))

        sheets_result.append({
            'sheet_name': sheet_name,
            'sheet_index': idx,
            'headers': headers,
            'rows_data': rows_data,
            'formulas': formulas,
            'merged_cells': merged,
            'row_count': len(rows_data),
            'col_count': max_col,
            'column_widths': col_widths
        })

    return sheets_result

def upload_document(body):
    file_data = body.get('file')
    file_name = body.get('file_name', 'document.xlsx')
    title = body.get('title', file_name)
    category = body.get('category', 'employee_registry')

    if not file_data:
        return json_response(400, {'error': 'Файл не передан'})

    if not file_name.lower().endswith(('.xlsx', '.xls')):
        return json_response(400, {'error': 'Поддерживаются только файлы .xlsx и .xls'})

    sheets = parse_excel(file_data)

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        INSERT INTO ohs_documents (title, category, file_name, sheets, metadata)
        VALUES ('%s', '%s', '%s', '%s', '{}')
        RETURNING id
    """ % (
        title.replace("'", "''"),
        category.replace("'", "''"),
        file_name.replace("'", "''"),
        json.dumps([s['sheet_name'] for s in sheets], ensure_ascii=False).replace("'", "''")
    ))
    doc_id = cur.fetchone()[0]

    for sheet in sheets:
        cur.execute("""
            INSERT INTO ohs_sheets (document_id, sheet_name, sheet_index, headers, rows_data, formulas, merged_cells, column_widths, row_count, col_count)
            VALUES (%d, '%s', %d, '%s', '%s', '%s', '%s', '%s', %d, %d)
        """ % (
            doc_id,
            sheet['sheet_name'].replace("'", "''"),
            sheet['sheet_index'],
            json.dumps(sheet['headers'], ensure_ascii=False).replace("'", "''"),
            json.dumps(sheet['rows_data'], ensure_ascii=False).replace("'", "''"),
            json.dumps(sheet['formulas'], ensure_ascii=False).replace("'", "''"),
            json.dumps(sheet['merged_cells'], ensure_ascii=False).replace("'", "''"),
            json.dumps(sheet['column_widths'], ensure_ascii=False).replace("'", "''"),
            sheet['row_count'],
            sheet['col_count']
        ))

    conn.commit()
    cur.close()
    conn.close()

    return json_response(200, {
        'success': True,
        'document_id': doc_id,
        'sheets_count': len(sheets),
        'total_rows': sum(s['row_count'] for s in sheets),
        'message': 'Документ загружен: %d листов, %d строк данных' % (len(sheets), sum(s['row_count'] for s in sheets))
    })

def get_documents(params):
    category = params.get('category', '')
    conn = get_db()
    cur = conn.cursor()

    where = ""
    if category:
        where = "WHERE d.category = '%s'" % category.replace("'", "''")

    cur.execute("""
        SELECT d.id, d.title, d.category, d.file_name, d.sheets, d.metadata, d.created_at, d.updated_at,
               COUNT(s.id) as sheets_count, COALESCE(SUM(s.row_count), 0) as total_rows
        FROM ohs_documents d
        LEFT JOIN ohs_sheets s ON s.document_id = d.id
        %s
        GROUP BY d.id
        ORDER BY d.created_at DESC
    """ % where)

    cols = [desc[0] for desc in cur.description]
    rows = [dict(zip(cols, row)) for row in cur.fetchall()]

    cur.close()
    conn.close()
    return json_response(200, {'documents': rows})

def get_document(params):
    doc_id = params.get('id', '')
    if not doc_id:
        return json_response(400, {'error': 'ID документа не указан'})

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT id, title, category, file_name, sheets, metadata, created_at, updated_at
        FROM ohs_documents WHERE id = %s
    """ % doc_id)
    doc_row = cur.fetchone()
    if not doc_row:
        cur.close()
        conn.close()
        return json_response(404, {'error': 'Документ не найден'})

    cols = [desc[0] for desc in cur.description]
    doc = dict(zip(cols, doc_row))

    cur.execute("""
        SELECT id, sheet_name, sheet_index, headers, rows_data, formulas, merged_cells, column_widths, row_count, col_count
        FROM ohs_sheets WHERE document_id = %s ORDER BY sheet_index
    """ % doc_id)
    sheet_cols = [desc[0] for desc in cur.description]
    sheets = [dict(zip(sheet_cols, row)) for row in cur.fetchall()]

    cur.close()
    conn.close()

    doc['sheets_data'] = sheets
    return json_response(200, doc)

def update_cell(body):
    doc_id = body.get('document_id')
    sheet_id = body.get('sheet_id')
    row_idx = body.get('row_index')
    col_idx = body.get('col_index')
    value = body.get('value', '')

    if doc_id is None or sheet_id is None or row_idx is None or col_idx is None:
        return json_response(400, {'error': 'Не указаны обязательные параметры'})

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT rows_data, formulas FROM ohs_sheets WHERE id = %d AND document_id = %d" % (sheet_id, doc_id))
    row = cur.fetchone()
    if not row:
        cur.close()
        conn.close()
        return json_response(404, {'error': 'Лист не найден'})

    rows_data = row[0] if isinstance(row[0], list) else json.loads(row[0])
    formulas = row[1] if isinstance(row[1], dict) else json.loads(row[1])

    if row_idx < len(rows_data) and col_idx < len(rows_data[row_idx]):
        rows_data[row_idx][col_idx] = value

    if isinstance(value, str) and value.startswith('='):
        cell_ref = '%s%d' % (col_letter(col_idx), row_idx + 2)
        formulas[cell_ref] = value
    else:
        cell_ref = '%s%d' % (col_letter(col_idx), row_idx + 2)
        if cell_ref in formulas:
            del formulas[cell_ref]

    cur.execute("""
        UPDATE ohs_sheets SET rows_data = '%s', formulas = '%s' WHERE id = %d
    """ % (
        json.dumps(rows_data, ensure_ascii=False).replace("'", "''"),
        json.dumps(formulas, ensure_ascii=False).replace("'", "''"),
        sheet_id
    ))

    cur.execute("UPDATE ohs_documents SET updated_at = NOW() WHERE id = %d" % doc_id)

    conn.commit()
    cur.close()
    conn.close()

    return json_response(200, {'success': True})

def delete_document(body):
    doc_id = body.get('document_id')
    if not doc_id:
        return json_response(400, {'error': 'ID документа не указан'})

    conn = get_db()
    cur = conn.cursor()

    cur.execute("UPDATE ohs_sheets SET rows_data = '[]', headers = '[]', formulas = '{}' WHERE document_id = %d" % doc_id)
    cur.execute("UPDATE ohs_documents SET title = '[удалён]', metadata = '{\"deleted\": true}' WHERE id = %d" % doc_id)

    conn.commit()
    cur.close()
    conn.close()
    return json_response(200, {'success': True})

def handler(event, context):
    """ОТ и ПБ — загрузка Excel-документов, парсинг листов с формулами, электронный реестр"""
    if event.get('httpMethod') == 'OPTIONS':
        return json_response(200, '')

    method = event.get('httpMethod', 'GET')
    params = event.get('queryStringParameters') or {}
    action = params.get('action', '')
    body = json.loads(event.get('body', '{}') or '{}')

    if method == 'POST' and action == 'upload':
        return upload_document(body)

    if method == 'GET' and action == 'documents':
        return get_documents(params)

    if method == 'GET' and action == 'document':
        return get_document(params)

    if method == 'PUT' and action == 'cell':
        return update_cell(body)

    if method == 'POST' and action == 'delete':
        return delete_document(body)

    return json_response(400, {'error': 'Неизвестное действие: %s' % action})
