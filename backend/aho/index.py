import json
import os
import base64
import uuid
from datetime import datetime, date as date_type
import psycopg2

def get_db():
    return psycopg2.connect(os.environ['DATABASE_URL'])

def serialize_default(obj):
    if isinstance(obj, datetime):
        if obj.tzinfo is None:
            return obj.isoformat() + '+00:00'
        return obj.isoformat()
    if isinstance(obj, (date_type,)):
        return obj.isoformat()
    return str(obj)

def json_response(status, body):
    return {
        'statusCode': status,
        'headers': {
            'Content-Type': 'application/json',
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Methods': 'GET, POST, PUT, DELETE, OPTIONS',
            'Access-Control-Allow-Headers': 'Content-Type, Authorization, X-Authorization'
        },
        'body': json.dumps(body, ensure_ascii=False, default=serialize_default)
    }

def handler(event, context):
    """АХО — загрузка списков, контроль въезда/выезда, расселение, статистика"""
    if event.get('httpMethod') == 'OPTIONS':
        return json_response(200, '')

    method = event.get('httpMethod', 'GET')
    params = event.get('queryStringParameters') or {}
    action = params.get('action', '')
    body = json.loads(event.get('body', '{}') or '{}')

    if method == 'POST' and action == 'upload':
        return upload_excel(body)
    elif method == 'GET' and action == 'list':
        return get_arrivals(params)
    elif method == 'GET' and action == 'batches':
        return get_batches(params)
    elif method == 'PUT' and action == 'checkin':
        return check_in(body)
    elif method == 'PUT' and action == 'checkout':
        return check_out(body)
    elif method == 'PUT' and action == 'assign-room':
        return assign_room(body)
    elif method == 'GET' and action == 'stats':
        return get_stats()
    elif method == 'GET' and action == 'medical-status':
        return get_medical_status(params)
    elif method == 'GET' and action == 'template':
        return get_template()
    elif method == 'PUT' and action == 'mass-checkin':
        return mass_check_in(body)
    elif method == 'PUT' and action == 'mass-checkout':
        return mass_check_out(body)
    elif method == 'GET' and action == 'medical-itr-stats':
        return get_medical_itr_stats(params)
    elif method == 'GET' and action == 'itr-positions':
        return get_itr_positions()
    elif method == 'PUT' and action == 'itr-positions':
        return save_itr_positions(body)
    elif method == 'POST' and action == 'reset':
        return perform_reset(body)
    elif method == 'GET' and action == 'export-all':
        return export_all_data()

    return json_response(404, {'error': 'Маршрут не найден'})


def parse_excel(file_bytes):
    import openpyxl
    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    rows_data = []
    headers = []
    header_map = {
        'фио': 'full_name', 'ф.и.о': 'full_name', 'ф.и.о.': 'full_name',
        'имя': 'full_name', 'фамилия': 'full_name', 'сотрудник': 'full_name',
        'должность': 'position', 'позиция': 'position',
        'подразделение': 'department', 'отдел': 'department', 'участок': 'department',
        'организация': 'organization', 'компания': 'organization', 'предприятие': 'organization',
        'телефон': 'phone', 'тел': 'phone', 'тел.': 'phone', 'номер': 'phone',
        'дата прибытия': 'arrival_date', 'дата приезда': 'arrival_date', 'заезд': 'arrival_date', 'прибытие': 'arrival_date',
        'дата отъезда': 'departure_date', 'дата убытия': 'departure_date', 'выезд': 'departure_date', 'отъезд': 'departure_date',
        'примечание': 'notes', 'примечания': 'notes', 'комментарий': 'notes',
    }

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            for cell in row:
                val = str(cell or '').strip().lower()
                mapped = header_map.get(val, '')
                headers.append(mapped)
            continue

        if all(c is None or str(c).strip() == '' for c in row):
            continue

        item = {}
        for j, cell in enumerate(row):
            if j < len(headers) and headers[j]:
                val = cell
                if headers[j] in ('arrival_date', 'departure_date'):
                    if isinstance(val, datetime):
                        val = val.date().isoformat()
                    elif isinstance(val, date_type):
                        val = val.isoformat()
                    elif val:
                        val = str(val).strip()
                    else:
                        val = None
                else:
                    val = str(val or '').strip()
                item[headers[j]] = val

        if item.get('full_name'):
            rows_data.append(item)

    wb.close()
    return rows_data


def generate_code(cur):
    cur.execute("SELECT COALESCE(MAX(id), 0) + 1 FROM personnel")
    next_id = cur.fetchone()[0]
    code = 'МК-%03d' % next_id
    qr = 'QR-MK-%03d' % next_id
    return code, qr


def upload_excel(body):
    file_data = body.get('file', '')
    file_name = body.get('file_name', 'upload.xlsx')
    arrival_date = body.get('arrival_date', '')
    departure_date = body.get('departure_date', '')
    org_type = body.get('organization_type', 'contractor')

    if not file_data:
        return json_response(400, {'error': 'Файл не передан'})

    try:
        file_bytes = base64.b64decode(file_data)
    except Exception:
        return json_response(400, {'error': 'Ошибка декодирования файла'})

    try:
        rows_data = parse_excel(file_bytes)
    except Exception as e:
        return json_response(400, {'error': 'Ошибка чтения Excel: ' + str(e)[:100]})

    if not rows_data:
        return json_response(400, {'error': 'Файл пустой или не распознан. Убедитесь, что в первой строке заголовки: ФИО, Должность, Подразделение и т.д.'})

    batch_id = 'AHO-' + datetime.now().strftime('%Y%m%d-%H%M%S') + '-' + uuid.uuid4().hex[:4]

    conn = get_db()
    cur = conn.cursor()

    created_ids = []
    for item in rows_data:
        full_name = item.get('full_name', '')
        if not full_name:
            continue

        position = item.get('position', '')
        department = item.get('department', '')
        organization = item.get('organization', '')
        phone = item.get('phone', '')
        notes = item.get('notes', '')
        item_arrival = item.get('arrival_date') or arrival_date or datetime.now().strftime('%Y-%m-%d')
        item_departure = item.get('departure_date') or departure_date or None

        personal_code, qr_code = generate_code(cur)

        cur.execute("""
            INSERT INTO personnel (personal_code, full_name, position, department, category, phone, status, qr_code, organization, organization_type, medical_status)
            VALUES ('%s', '%s', '%s', '%s', '%s', '%s', 'expected', '%s', '%s', '%s', 'pending')
            RETURNING id
        """ % (
            personal_code,
            full_name.replace("'", "''"),
            position.replace("'", "''"),
            department.replace("'", "''"),
            org_type.replace("'", "''"),
            phone.replace("'", "''"),
            qr_code,
            organization.replace("'", "''"),
            org_type.replace("'", "''"),
        ))
        person_id = cur.fetchone()[0]

        dep_sql = "NULL"
        if item_departure:
            dep_sql = "'%s'" % str(item_departure).replace("'", "''")

        cur.execute("""
            INSERT INTO aho_arrivals (batch_id, personnel_id, full_name, position, department, organization, organization_type, phone, arrival_date, departure_date, personal_code, notes, arrival_status)
            VALUES ('%s', %d, '%s', '%s', '%s', '%s', '%s', '%s', '%s', %s, '%s', '%s', 'expected')
            RETURNING id
        """ % (
            batch_id,
            person_id,
            full_name.replace("'", "''"),
            position.replace("'", "''"),
            department.replace("'", "''"),
            organization.replace("'", "''"),
            org_type.replace("'", "''"),
            phone.replace("'", "''"),
            str(item_arrival).replace("'", "''"),
            dep_sql,
            personal_code,
            notes.replace("'", "''"),
        ))
        aho_id = cur.fetchone()[0]
        created_ids.append({'aho_id': aho_id, 'person_id': person_id, 'personal_code': personal_code, 'qr_code': qr_code, 'full_name': full_name})

        cur.execute("""
            INSERT INTO events (event_type, description, personnel_id)
            VALUES ('aho_upload', 'Загружен через АХО (партия %s): %s', %d)
        """ % (batch_id, full_name.replace("'", "''"), person_id))

    cur.execute("""
        INSERT INTO aho_batches (batch_id, file_name, total_count, arrival_date, departure_date)
        VALUES ('%s', '%s', %d, %s, %s)
    """ % (
        batch_id,
        file_name.replace("'", "''"),
        len(created_ids),
        ("'%s'" % arrival_date) if arrival_date else 'NULL',
        ("'%s'" % departure_date) if departure_date else 'NULL',
    ))

    conn.commit()
    cur.close()
    conn.close()

    return json_response(200, {
        'message': 'Загружено %d человек' % len(created_ids),
        'batch_id': batch_id,
        'total': len(created_ids),
        'persons': created_ids
    })


def get_arrivals(params):
    conn = get_db()
    cur = conn.cursor()

    batch_id = params.get('batch_id', '')
    status = params.get('status', '')
    date_from = params.get('date_from', '')
    date_to = params.get('date_to', '')

    query = """
        SELECT a.id, a.batch_id, a.personnel_id, a.full_name, a.position, a.department,
               a.organization, a.organization_type, a.phone, a.arrival_date, a.departure_date,
               a.arrival_status, a.check_in_at, a.check_out_at, a.room, a.building,
               a.medical_status, a.personal_code, a.notes, a.created_at,
               p.medical_status as current_medical, p.status as person_status, p.qr_code
        FROM aho_arrivals a
        LEFT JOIN personnel p ON a.personnel_id = p.id
        WHERE a.is_hidden = FALSE
    """
    if batch_id:
        query += " AND a.batch_id = '%s'" % batch_id.replace("'", "''")
    if status:
        query += " AND a.arrival_status = '%s'" % status.replace("'", "''")
    if date_from:
        query += " AND a.arrival_date >= '%s'" % date_from.replace("'", "''")
    if date_to:
        query += " AND a.arrival_date <= '%s'" % date_to.replace("'", "''")
    query += " ORDER BY a.arrival_date DESC, a.full_name ASC LIMIT 500"

    cur.execute(query)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    items = []
    for r in rows:
        items.append({
            'id': r[0], 'batch_id': r[1], 'personnel_id': r[2],
            'full_name': r[3], 'position': r[4], 'department': r[5],
            'organization': r[6], 'organization_type': r[7], 'phone': r[8],
            'arrival_date': r[9], 'departure_date': r[10],
            'arrival_status': r[11], 'check_in_at': r[12], 'check_out_at': r[13],
            'room': r[14], 'building': r[15],
            'medical_status': r[16], 'personal_code': r[17], 'notes': r[18],
            'created_at': r[19],
            'current_medical': r[20], 'person_status': r[21], 'qr_code': r[22],
        })

    return json_response(200, {'items': items, 'total': len(items)})


def get_batches(params):
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        SELECT b.id, b.batch_id, b.file_name, b.total_count, b.arrived_count, b.departed_count,
               b.arrival_date, b.departure_date, b.created_at,
               (SELECT COUNT(*) FROM aho_arrivals WHERE batch_id = b.batch_id AND arrival_status = 'arrived') as real_arrived,
               (SELECT COUNT(*) FROM aho_arrivals WHERE batch_id = b.batch_id AND arrival_status = 'departed') as real_departed
        FROM aho_batches b
        WHERE b.is_hidden = FALSE
        ORDER BY b.created_at DESC
        LIMIT 50
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    batches = []
    for r in rows:
        batches.append({
            'id': r[0], 'batch_id': r[1], 'file_name': r[2],
            'total_count': r[3], 'arrived_count': r[9], 'departed_count': r[10],
            'arrival_date': r[6], 'departure_date': r[7], 'created_at': r[8],
        })

    return json_response(200, {'batches': batches})


def check_in(body):
    arrival_id = body.get('id')
    if not arrival_id:
        return json_response(400, {'error': 'ID записи обязателен'})

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        UPDATE aho_arrivals SET arrival_status = 'arrived', check_in_at = NOW(), updated_at = NOW()
        WHERE id = %d RETURNING personnel_id, full_name
    """ % int(arrival_id))
    row = cur.fetchone()

    if not row:
        cur.close()
        conn.close()
        return json_response(404, {'error': 'Запись не найдена'})

    if row[0]:
        cur.execute("""
            UPDATE personnel SET status = 'arrived', updated_at = NOW() WHERE id = %d
        """ % row[0])
        cur.execute("""
            INSERT INTO events (event_type, description, personnel_id)
            VALUES ('check_in', '%s — въехал на территорию (АХО)', %d)
        """ % (row[1].replace("'", "''"), row[0]))

    conn.commit()
    cur.close()
    conn.close()

    return json_response(200, {'message': '%s — въезд зафиксирован' % row[1]})


def check_out(body):
    arrival_id = body.get('id')
    if not arrival_id:
        return json_response(400, {'error': 'ID записи обязателен'})

    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
        UPDATE aho_arrivals SET arrival_status = 'departed', check_out_at = NOW(), updated_at = NOW()
        WHERE id = %d RETURNING personnel_id, full_name
    """ % int(arrival_id))
    row = cur.fetchone()

    if not row:
        cur.close()
        conn.close()
        return json_response(404, {'error': 'Запись не найдена'})

    if row[0]:
        cur.execute("""
            UPDATE personnel SET status = 'departed', updated_at = NOW() WHERE id = %d
        """ % row[0])
        cur.execute("""
            INSERT INTO events (event_type, description, personnel_id)
            VALUES ('check_out', '%s — выехал с территории (АХО)', %d)
        """ % (row[1].replace("'", "''"), row[0]))

    conn.commit()
    cur.close()
    conn.close()

    return json_response(200, {'message': '%s — выезд зафиксирован' % row[1]})


def assign_room(body):
    arrival_id = body.get('id')
    room = body.get('room', '')
    building = body.get('building', '')

    if not arrival_id or not room:
        return json_response(400, {'error': 'ID и номер комнаты обязательны'})

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT id, capacity, occupied FROM rooms WHERE room_number = '%s'" % room.replace("'", "''"))
    room_row = cur.fetchone()
    if room_row and room_row[2] >= room_row[1]:
        cur.close()
        conn.close()
        return json_response(400, {'error': 'Комната %s заполнена (%d/%d)' % (room, room_row[2], room_row[1])})

    cur.execute("""
        UPDATE aho_arrivals SET room = '%s', building = '%s', updated_at = NOW()
        WHERE id = %d RETURNING personnel_id, full_name
    """ % (room.replace("'", "''"), building.replace("'", "''"), int(arrival_id)))
    row = cur.fetchone()

    if not row:
        cur.close()
        conn.close()
        return json_response(404, {'error': 'Запись не найдена'})

    if row[0]:
        cur.execute("""
            UPDATE personnel SET room = '%s', updated_at = NOW() WHERE id = %d
        """ % (room.replace("'", "''"), row[0]))

    if room_row:
        cur.execute("UPDATE rooms SET occupied = occupied + 1 WHERE id = %d" % room_row[0])

    cur.execute("""
        INSERT INTO events (event_type, description, personnel_id)
        VALUES ('housing', '%s — заселён в комнату %s', %s)
    """ % (row[1].replace("'", "''"), room.replace("'", "''"), row[0] if row[0] else 'NULL'))

    conn.commit()
    cur.close()
    conn.close()

    return json_response(200, {'message': '%s заселён в комнату %s' % (row[1], room)})


def get_stats():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM aho_arrivals WHERE is_hidden = FALSE")
    total = cur.fetchone()[0]

    cur.execute("SELECT arrival_status, COUNT(*) FROM aho_arrivals WHERE is_hidden = FALSE GROUP BY arrival_status")
    by_status = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("SELECT COUNT(*) FROM aho_arrivals WHERE is_hidden = FALSE AND room IS NOT NULL AND room != ''")
    housed = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM aho_arrivals WHERE is_hidden = FALSE AND (room IS NULL OR room = '')")
    not_housed = cur.fetchone()[0]

    cur.execute("""
        SELECT p.medical_status, COUNT(*)
        FROM aho_arrivals a
        JOIN personnel p ON a.personnel_id = p.id
        WHERE a.arrival_status = 'arrived' AND a.is_hidden = FALSE
        GROUP BY p.medical_status
    """)
    medical = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("""
        SELECT COUNT(*) FROM aho_arrivals
        WHERE is_hidden = FALSE AND arrival_date = CURRENT_DATE
    """)
    today_expected = cur.fetchone()[0]

    cur.execute("""
        SELECT COUNT(*) FROM aho_arrivals
        WHERE is_hidden = FALSE AND departure_date = CURRENT_DATE
    """)
    today_departing = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM aho_batches WHERE is_hidden = FALSE")
    total_batches = cur.fetchone()[0]

    cur.execute("SELECT COALESCE(SUM(capacity), 0), COALESCE(SUM(occupied), 0) FROM rooms")
    rooms_row = cur.fetchone()

    cur.close()
    conn.close()

    return json_response(200, {
        'total': total,
        'by_status': by_status,
        'housed': housed,
        'not_housed': not_housed,
        'medical': medical,
        'today_expected': today_expected,
        'today_departing': today_departing,
        'total_batches': total_batches,
        'rooms_capacity': rooms_row[0],
        'rooms_occupied': rooms_row[1],
    })


def get_medical_status(params):
    batch_id = params.get('batch_id', '')
    status_filter = params.get('medical', '')

    conn = get_db()
    cur = conn.cursor()

    query = """
        SELECT a.id, a.full_name, a.personal_code, a.department, a.organization,
               a.arrival_status, a.room,
               p.medical_status, p.id as person_id,
               mc.status as last_check_status, mc.checked_at as last_check_time,
               mc.blood_pressure, mc.pulse, mc.temperature, mc.alcohol_level, mc.notes as check_notes
        FROM aho_arrivals a
        LEFT JOIN personnel p ON a.personnel_id = p.id
        LEFT JOIN LATERAL (
            SELECT status, checked_at, blood_pressure, pulse, temperature, alcohol_level, notes
            FROM medical_checks
            WHERE personnel_id = p.id
            ORDER BY checked_at DESC
            LIMIT 1
        ) mc ON true
        WHERE a.arrival_status IN ('arrived', 'expected') AND a.is_hidden = FALSE
    """
    if batch_id:
        query += " AND a.batch_id = '%s'" % batch_id.replace("'", "''")
    if status_filter:
        query += " AND COALESCE(p.medical_status, 'pending') = '%s'" % status_filter.replace("'", "''")
    query += " ORDER BY p.medical_status ASC, a.full_name ASC LIMIT 300"

    cur.execute(query)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    items = []
    for r in rows:
        items.append({
            'id': r[0], 'full_name': r[1], 'personal_code': r[2],
            'department': r[3], 'organization': r[4],
            'arrival_status': r[5], 'room': r[6],
            'medical_status': r[7] or 'pending', 'person_id': r[8],
            'last_check_status': r[9], 'last_check_time': r[10],
            'blood_pressure': r[11], 'pulse': r[12],
            'temperature': str(r[13]) if r[13] else None,
            'alcohol_level': str(r[14]) if r[14] else None,
            'check_notes': r[15],
        })

    return json_response(200, {'items': items, 'total': len(items)})


def get_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Список въезжающих"

    headers = ["ФИО", "Должность", "Подразделение", "Организация", "Телефон", "Дата прибытия", "Дата отъезда", "Примечание"]
    col_widths = [35, 25, 25, 30, 18, 16, 16, 30]

    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2D5A27", end_color="2D5A27", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    data_font = Font(name="Arial", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    examples = [
        ["Иванов Иван Иванович", "Горный мастер", "Участок №1", "ООО Рудник", "+7 900 123-45-67", "2026-03-01", "2026-03-15", ""],
        ["Петров Пётр Петрович", "Электрослесарь", "Энергоцех", "ООО Рудник", "+7 900 765-43-21", "2026-03-01", "2026-03-15", ""],
        ["Сидорова Анна Сергеевна", "Инженер ОТ", "ОТиПБ", "ООО Подрядчик", "+7 900 111-22-33", "2026-03-01", "2026-03-08", "Командировка"],
    ]

    example_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    for row_idx, row_data in enumerate(examples, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10, color="999999", italic=True)
            cell.fill = example_fill
            cell.alignment = data_align
            cell.border = thin_border

    for row_idx in range(5, 55):
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx, value="")
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    ws_help = wb.create_sheet("Инструкция")
    ws_help.column_dimensions["A"].width = 80

    instructions = [
        "ИНСТРУКЦИЯ ПО ЗАПОЛНЕНИЮ",
        "",
        "1. Заполните список на листе «Список въезжающих»",
        "2. Первая строка (зелёная) — заголовки, НЕ удаляйте и НЕ меняйте их",
        "3. Строки 2-4 (серые) — примеры, можно удалить или заменить",
        "4. Начинайте заполнение с 5 строки",
        "",
        "ОБЯЗАТЕЛЬНЫЕ ПОЛЯ:",
        "• ФИО — полное имя сотрудника",
        "",
        "НЕОБЯЗАТЕЛЬНЫЕ ПОЛЯ:",
        "• Должность, Подразделение, Организация, Телефон",
        "• Дата прибытия/отъезда — формат ГГГГ-ММ-ДД (можно указать при загрузке)",
        "• Примечание — любая дополнительная информация",
        "",
        "ПОСЛЕ ЗАГРУЗКИ:",
        "• Каждому сотруднику автоматически присваивается личный код (МК-XXX)",
        "• Автоматически генерируется QR-код",
        "• Сотрудник появляется в разделе «Персонал»",
        "• Можно отслеживать въезд/выезд и медосмотр в разделе «АХО»",
    ]
    for i, text in enumerate(instructions, 1):
        cell = ws_help.cell(row=i, column=1, value=text)
        if i == 1:
            cell.font = Font(name="Arial", bold=True, size=14)
        elif text.startswith("ОБЯЗАТЕЛЬНЫЕ") or text.startswith("НЕОБЯЗАТЕЛЬНЫЕ") or text.startswith("ПОСЛЕ"):
            cell.font = Font(name="Arial", bold=True, size=11)
        else:
            cell.font = Font(name="Arial", size=10)

    buf = BytesIO()
    wb.save(buf)
    wb.close()
    file_b64 = base64.b64encode(buf.getvalue()).decode()

    return json_response(200, {
        'file': file_b64,
        'file_name': 'Шаблон_списка_въезжающих.xlsx'
    })


def mass_check_in(body):
    ids = body.get('ids', [])
    batch_id = body.get('batch_id', '')
    if not ids and not batch_id:
        return json_response(400, {'error': 'Укажите список ID или партию'})

    conn = get_db()
    cur = conn.cursor()

    if batch_id and not ids:
        cur.execute("SELECT id FROM aho_arrivals WHERE batch_id = '%s' AND arrival_status = 'expected'" % batch_id.replace("'", "''"))
        ids = [r[0] for r in cur.fetchall()]

    count = 0
    for aid in ids:
        cur.execute("""
            UPDATE aho_arrivals SET arrival_status = 'arrived', check_in_at = NOW(), updated_at = NOW()
            WHERE id = %d AND arrival_status = 'expected' RETURNING personnel_id, full_name
        """ % int(aid))
        row = cur.fetchone()
        if row:
            count += 1
            if row[0]:
                cur.execute("UPDATE personnel SET status = 'arrived', updated_at = NOW() WHERE id = %d" % row[0])
                cur.execute("""
                    INSERT INTO events (event_type, description, personnel_id)
                    VALUES ('check_in', '%s — массовый въезд (АХО)', %d)
                """ % (row[1].replace("'", "''"), row[0]))

    conn.commit()
    cur.close()
    conn.close()
    return json_response(200, {'message': 'Въезд зафиксирован: %d чел.' % count, 'count': count})


def mass_check_out(body):
    ids = body.get('ids', [])
    batch_id = body.get('batch_id', '')
    if not ids and not batch_id:
        return json_response(400, {'error': 'Укажите список ID или партию'})

    conn = get_db()
    cur = conn.cursor()

    if batch_id and not ids:
        cur.execute("SELECT id FROM aho_arrivals WHERE batch_id = '%s' AND arrival_status = 'arrived'" % batch_id.replace("'", "''"))
        ids = [r[0] for r in cur.fetchall()]

    count = 0
    for aid in ids:
        cur.execute("""
            UPDATE aho_arrivals SET arrival_status = 'departed', check_out_at = NOW(), updated_at = NOW()
            WHERE id = %d AND arrival_status = 'arrived' RETURNING personnel_id, full_name
        """ % int(aid))
        row = cur.fetchone()
        if row:
            count += 1
            if row[0]:
                cur.execute("UPDATE personnel SET status = 'departed', updated_at = NOW() WHERE id = %d" % row[0])
                cur.execute("""
                    INSERT INTO events (event_type, description, personnel_id)
                    VALUES ('check_out', '%s — массовый выезд (АХО)', %d)
                """ % (row[1].replace("'", "''"), row[0]))

    conn.commit()
    cur.close()
    conn.close()
    return json_response(200, {'message': 'Выезд зафиксирован: %d чел.' % count, 'count': count})


def get_itr_positions():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT value FROM settings WHERE key = 'itr_positions'")
    row = cur.fetchone()
    cur.close()
    conn.close()
    positions = row[0] if row else []
    return json_response(200, {'positions': positions})


def save_itr_positions(body):
    positions = body.get('positions', [])
    if not isinstance(positions, list):
        return json_response(400, {'error': 'positions должен быть массивом'})

    conn = get_db()
    cur = conn.cursor()
    positions_json = json.dumps(positions, ensure_ascii=False)
    cur.execute("""
        INSERT INTO settings (key, value, updated_at) VALUES ('itr_positions', '%s'::jsonb, NOW())
        ON CONFLICT (key) DO UPDATE SET value = '%s'::jsonb, updated_at = NOW()
    """ % (positions_json.replace("'", "''"), positions_json.replace("'", "''")))
    conn.commit()
    cur.close()
    conn.close()
    return json_response(200, {'message': 'Список ИТР должностей сохранён', 'count': len(positions)})


def get_medical_itr_stats(params):
    batch_id = params.get('batch_id', '')

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT value FROM settings WHERE key = 'itr_positions'")
    row = cur.fetchone()
    itr_positions = row[0] if row else []

    itr_conditions = []
    for pos in itr_positions:
        safe_pos = pos.replace("'", "''").lower()
        itr_conditions.append("LOWER(COALESCE(a.position, '')) LIKE '%%%s%%'" % safe_pos)

    if itr_conditions:
        itr_where = "(%s)" % " OR ".join(itr_conditions)
    else:
        itr_where = "FALSE"

    base_where = "a.arrival_status IN ('arrived', 'expected') AND a.is_hidden = FALSE"
    if batch_id:
        base_where += " AND a.batch_id = '%s'" % batch_id.replace("'", "''")

    cur.execute("""
        SELECT
            COUNT(*) FILTER (WHERE %s) as itr_total,
            COUNT(*) FILTER (WHERE NOT (%s)) as worker_total,
            COUNT(*) FILTER (WHERE %s AND COALESCE(p.medical_status, 'pending') = 'passed') as itr_passed,
            COUNT(*) FILTER (WHERE %s AND COALESCE(p.medical_status, 'pending') = 'failed') as itr_failed,
            COUNT(*) FILTER (WHERE %s AND COALESCE(p.medical_status, 'pending') = 'pending') as itr_pending,
            COUNT(*) FILTER (WHERE NOT (%s) AND COALESCE(p.medical_status, 'pending') = 'passed') as worker_passed,
            COUNT(*) FILTER (WHERE NOT (%s) AND COALESCE(p.medical_status, 'pending') = 'failed') as worker_failed,
            COUNT(*) FILTER (WHERE NOT (%s) AND COALESCE(p.medical_status, 'pending') = 'pending') as worker_pending,
            COUNT(*) as total
        FROM aho_arrivals a
        LEFT JOIN personnel p ON a.personnel_id = p.id
        WHERE %s
    """ % (itr_where, itr_where, itr_where, itr_where, itr_where, itr_where, itr_where, itr_where, base_where))
    r = cur.fetchone()

    itr_list_query = """
        SELECT a.id, a.full_name, a.personal_code, a.position, a.department, a.organization,
               COALESCE(p.medical_status, 'pending') as medical_status,
               mc.status as last_check, mc.checked_at, mc.blood_pressure, mc.pulse, mc.temperature, mc.alcohol_level
        FROM aho_arrivals a
        LEFT JOIN personnel p ON a.personnel_id = p.id
        LEFT JOIN LATERAL (
            SELECT status, checked_at, blood_pressure, pulse, temperature, alcohol_level
            FROM medical_checks WHERE personnel_id = p.id ORDER BY checked_at DESC LIMIT 1
        ) mc ON true
        WHERE %s AND %s
        ORDER BY p.medical_status ASC, a.full_name ASC
    """ % (base_where, itr_where)
    cur.execute(itr_list_query)
    itr_rows = cur.fetchall()

    worker_list_query = """
        SELECT a.id, a.full_name, a.personal_code, a.position, a.department, a.organization,
               COALESCE(p.medical_status, 'pending') as medical_status,
               mc.status as last_check, mc.checked_at, mc.blood_pressure, mc.pulse, mc.temperature, mc.alcohol_level
        FROM aho_arrivals a
        LEFT JOIN personnel p ON a.personnel_id = p.id
        LEFT JOIN LATERAL (
            SELECT status, checked_at, blood_pressure, pulse, temperature, alcohol_level
            FROM medical_checks WHERE personnel_id = p.id ORDER BY checked_at DESC LIMIT 1
        ) mc ON true
        WHERE %s AND NOT (%s)
        ORDER BY p.medical_status ASC, a.full_name ASC
    """ % (base_where, itr_where)
    cur.execute(worker_list_query)
    worker_rows = cur.fetchall()

    cur.close()
    conn.close()

    def map_rows(rows):
        items = []
        for row in rows:
            items.append({
                'id': row[0], 'full_name': row[1], 'personal_code': row[2],
                'position': row[3], 'department': row[4], 'organization': row[5],
                'medical_status': row[6],
                'last_check': row[7], 'checked_at': row[8],
                'blood_pressure': row[9], 'pulse': row[10],
                'temperature': str(row[11]) if row[11] else None,
                'alcohol_level': str(row[12]) if row[12] else None,
            })
        return items

    return json_response(200, {
        'summary': {
            'itr_total': r[0], 'worker_total': r[1],
            'itr_passed': r[2], 'itr_failed': r[3], 'itr_pending': r[4],
            'worker_passed': r[5], 'worker_failed': r[6], 'worker_pending': r[7],
            'total': r[8],
        },
        'itr_list': map_rows(itr_rows),
        'worker_list': map_rows(worker_rows),
        'itr_positions': itr_positions,
    })


def perform_reset(body):
    """Обнуление данных системы. Soft-delete скрывает с экрана, full — удаляет из БД"""
    reset_type = body.get('reset_type', '')
    valid_types = ['personnel', 'aho_arrivals', 'aho_departures', 'delete_aho_arrivals', 'delete_aho_departures', 'medical', 'lamp_room', 'full']
    if reset_type not in valid_types:
        return json_response(400, {'error': 'Неизвестный тип обнуления. Допустимые: %s' % ', '.join(valid_types)})

    conn = get_db()
    cur = conn.cursor()
    affected = 0

    if reset_type == 'personnel':
        cur.execute("""
            UPDATE personnel SET is_hidden = TRUE, updated_at = NOW()
            WHERE is_hidden = FALSE AND personal_code NOT IN (
                SELECT COALESCE(p.personal_code, '') FROM users u
                LEFT JOIN personnel p ON p.full_name = u.full_name
                WHERE u.role = 'admin' AND p.personal_code IS NOT NULL
            )
        """)
        affected = cur.rowcount
        cur.execute("""
            INSERT INTO reset_log (reset_type, description, affected_rows)
            VALUES ('personnel', 'Обнуление списка персонала (скрыто %d записей)', %d)
        """ % (affected, affected))
        cur.execute("""
            INSERT INTO events (event_type, description)
            VALUES ('system_reset', 'Обнуление списка персонала — скрыто %d записей')
        """ % affected)

    elif reset_type == 'aho_arrivals':
        cur.execute("""
            UPDATE aho_arrivals SET is_hidden = TRUE, updated_at = NOW()
            WHERE is_hidden = FALSE AND arrival_status IN ('expected', 'arrived')
        """)
        affected = cur.rowcount
        cur.execute("""
            UPDATE aho_batches SET is_hidden = TRUE WHERE is_hidden = FALSE
        """)
        cur.execute("""
            INSERT INTO reset_log (reset_type, description, affected_rows)
            VALUES ('aho_arrivals', 'Обнуление списка заехавших АХО (скрыто %d записей)', %d)
        """ % (affected, affected))
        cur.execute("""
            INSERT INTO events (event_type, description)
            VALUES ('system_reset', 'Обнуление списка заехавших АХО — скрыто %d записей')
        """ % affected)

    elif reset_type == 'aho_departures':
        cur.execute("""
            UPDATE aho_arrivals SET is_hidden = TRUE, updated_at = NOW()
            WHERE is_hidden = FALSE AND arrival_status = 'departed'
        """)
        affected = cur.rowcount
        cur.execute("""
            INSERT INTO reset_log (reset_type, description, affected_rows)
            VALUES ('aho_departures', 'Обнуление списка выехавших АХО (скрыто %d записей)', %d)
        """ % (affected, affected))
        cur.execute("""
            INSERT INTO events (event_type, description)
            VALUES ('system_reset', 'Обнуление списка выехавших АХО — скрыто %d записей')
        """ % affected)

    elif reset_type == 'delete_aho_arrivals':
        cur.execute("""
            SELECT COUNT(*) FROM aho_arrivals WHERE arrival_status IN ('expected', 'arrived')
        """)
        affected = cur.fetchone()[0]
        cur.execute("DELETE FROM aho_arrivals WHERE arrival_status IN ('expected', 'arrived')")
        cur.execute("DELETE FROM aho_batches WHERE batch_id NOT IN (SELECT DISTINCT batch_id FROM aho_arrivals)")
        cur.execute("""
            INSERT INTO reset_log (reset_type, description, affected_rows)
            VALUES ('delete_aho_arrivals', 'Удаление списка заехавших АХО (удалено %d записей)', %d)
        """ % (affected, affected))
        cur.execute("""
            INSERT INTO events (event_type, description)
            VALUES ('system_reset', 'Удаление списка заехавших АХО — удалено %d записей из базы')
        """ % affected)

    elif reset_type == 'delete_aho_departures':
        cur.execute("""
            SELECT COUNT(*) FROM aho_arrivals WHERE arrival_status = 'departed'
        """)
        affected = cur.fetchone()[0]
        cur.execute("DELETE FROM aho_arrivals WHERE arrival_status = 'departed'")
        cur.execute("DELETE FROM aho_batches WHERE batch_id NOT IN (SELECT DISTINCT batch_id FROM aho_arrivals)")
        cur.execute("""
            INSERT INTO reset_log (reset_type, description, affected_rows)
            VALUES ('delete_aho_departures', 'Удаление списка выехавших АХО (удалено %d записей)', %d)
        """ % (affected, affected))
        cur.execute("""
            INSERT INTO events (event_type, description)
            VALUES ('system_reset', 'Удаление списка выехавших АХО — удалено %d записей из базы')
        """ % affected)

    elif reset_type == 'medical':
        cur.execute("""
            UPDATE medical_checks SET is_hidden = TRUE
            WHERE is_hidden = FALSE
        """)
        affected = cur.rowcount
        cur.execute("""
            UPDATE personnel SET medical_status = 'pending', updated_at = NOW()
            WHERE medical_status != 'pending' AND status != 'archived'
        """)
        cur.execute("""
            INSERT INTO reset_log (reset_type, description, affected_rows)
            VALUES ('medical', 'Обнуление списков медосмотров (скрыто %d записей)', %d)
        """ % (affected, affected))
        cur.execute("""
            INSERT INTO events (event_type, description)
            VALUES ('system_reset', 'Обнуление медосмотров — скрыто %d записей')
        """ % affected)

    elif reset_type == 'lamp_room':
        affected = 0
        for table in ['lamp_room_issues', 'lamp_room_denials']:
            cur.execute("SELECT COUNT(*) FROM %s" % table)
            affected += cur.fetchone()[0]
            cur.execute("DELETE FROM %s" % table)
        cur.execute("UPDATE lamp_room_equipment SET status = 'available', notes = NULL WHERE status = 'issued'")
        cur.execute("""
            INSERT INTO reset_log (reset_type, description, affected_rows)
            VALUES ('lamp_room', 'Обнуление ламповой — удалено %d записей выдачи/приёма', %d)
        """ % (affected, affected))
        cur.execute("""
            INSERT INTO events (event_type, description)
            VALUES ('system_reset', 'Обнуление ламповой — удалено %d записей выдачи и приёма фонарей/самоспасателей')
        """ % affected)

    elif reset_type == 'full':
        counts = {}
        for table in ['aho_arrivals', 'aho_batches', 'medical_checks', 'events', 'notifications',
                       'dispatcher_messages', 'medical_reset_log', 'lanterns', 'rooms',
                       'lamp_room_issues', 'lamp_room_denials',
                       'security_checks', 'checkpoint_passes']:
            cur.execute("SELECT COUNT(*) FROM %s" % table)
            counts[table] = cur.fetchone()[0]
            cur.execute("DELETE FROM %s" % table)

        cur.execute("SELECT COUNT(*) FROM personnel WHERE full_name NOT IN (SELECT full_name FROM users WHERE role = 'admin')")
        counts['personnel'] = cur.fetchone()[0]
        cur.execute("DELETE FROM sessions WHERE user_id NOT IN (SELECT id FROM users WHERE role = 'admin')")
        cur.execute("DELETE FROM personnel WHERE full_name NOT IN (SELECT full_name FROM users WHERE role = 'admin')")
        cur.execute("DELETE FROM users WHERE role != 'admin'")

        affected = sum(counts.values())
        cur.execute("""
            INSERT INTO reset_log (reset_type, description, affected_rows)
            VALUES ('full', 'Полный сброс системы. Администраторы сохранены. Удалено: %d записей', %d)
        """ % (affected, affected))
        cur.execute("""
            INSERT INTO events (event_type, description)
            VALUES ('system_reset', 'Полный сброс системы — удалено %d записей. Администраторы сохранены.')
        """ % affected)

    conn.commit()
    cur.close()
    conn.close()

    labels = {
        'personnel': 'Список персонала обнулён',
        'aho_arrivals': 'Список заехавших по АХО обнулён',
        'aho_departures': 'Список выехавших по АХО обнулён',
        'delete_aho_arrivals': 'Список заехавших по АХО удалён из базы',
        'delete_aho_departures': 'Список выехавших по АХО удалён из базы',
        'medical': 'Списки медосмотров обнулены',
        'lamp_room': 'Списки ламповой обнулены',
        'full': 'Система полностью сброшена до заводских настроек',
    }

    return json_response(200, {
        'message': labels.get(reset_type, 'Готово'),
        'affected': affected,
        'reset_type': reset_type,
    })


def export_all_data():
    """Выгрузка всех данных из БД (включая скрытые) в JSON"""
    conn = get_db()
    cur = conn.cursor()

    tables_cols = {
        'personnel': ['id','personal_code','full_name','position','department','category','phone','room','status','qr_code','medical_status','shift','organization','organization_type','is_hidden','created_at','updated_at'],
        'aho_arrivals': ['id','batch_id','personnel_id','full_name','position','department','organization','organization_type','phone','arrival_date','departure_date','arrival_status','check_in_at','check_out_at','room','building','medical_status','personal_code','notes','is_hidden','created_at','updated_at'],
        'aho_batches': ['id','batch_id','file_name','total_count','arrived_count','departed_count','arrival_date','departure_date','is_hidden','created_at'],
        'medical_checks': ['id','personnel_id','check_type','status','blood_pressure','pulse','temperature','alcohol_level','shift_type','check_direction','shift_date','doctor_name','notes','is_hidden','checked_at'],
        'events': ['id','event_type','description','personnel_id','user_id','metadata','is_hidden','created_at'],
        'lanterns': ['id','lantern_number','rescuer_number','status','assigned_to','issued_at','returned_at','condition'],
        'rooms': ['id','room_number','building','capacity','occupied','status'],
        'notifications': ['id','type','title','message','person_name','person_code','is_read','created_at'],
        'reset_log': ['id','reset_type','description','affected_rows','performed_by','performed_at'],
    }

    result = {}
    for table, cols in tables_cols.items():
        cur.execute("SELECT %s FROM %s ORDER BY id" % (', '.join(cols), table))
        rows = cur.fetchall()
        result[table] = []
        for r in rows:
            item = {}
            for i, col in enumerate(cols):
                item[col] = r[i]
            result[table].append(item)
        result[table + '_count'] = len(rows)

    cur.close()
    conn.close()

    return json_response(200, {
        'data': result,
        'exported_at': datetime.now().isoformat(),
        'message': 'Полная выгрузка всех данных (включая скрытые)',
    })